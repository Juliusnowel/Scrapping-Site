import os
import re
import json
import time
import logging
import zipfile
from urllib.parse import urlparse, urljoin, urlunparse, parse_qs, urlencode

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

SHOPEE_PRODUCT_RE = re.compile(
    r"/product/\d+/\d+|-i\.\d+\.\d+|/[^/]*-i\.\d+\.\d+",
    re.I,
)
LAZADA_PRODUCT_RE = re.compile(r"-i\d+.*\.html", re.I)
SHOPEE_ITEM_API = "https://shopee.ph/api/v4/item/get"

# ----------------- helpers -----------------

def parse_shopee_ids(url: str) -> tuple[int, int] | None:
    """
    Extract (shopid, itemid) from ...-i.<shopid>.<itemid> product URLs.
    """
    path = urlparse(url).path
    m = re.search(r"-i\.(\d+)\.(\d+)", path or "")
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))

def slugify(text: str, max_len: int = 60) -> str:
    text = (text or "").strip()
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text)
    text = text.strip("-").lower()
    return text[:max_len] or "item"


def is_shopee_product_url(url: str) -> bool:
    path = urlparse(url).path
    return SHOPEE_PRODUCT_RE.search(path or "") is not None

def get_platform(url: str) -> str:
    host = urlparse(url).netloc.lower()
    if "shopee" in host:
        return "shopee"
    if "lazada" in host:
        return "lazada"
    return "unknown"


def fetch(url: str, timeout: int = 20) -> requests.Response:
    logging.info("GET %s", url)
    r = SESSION.get(url, timeout=timeout)
    r.raise_for_status()
    return r


def fetch_rendered_html_with_browser(url: str, timeout_ms: int = 30000) -> str:
    """
    Use Playwright to render JS-heavy pages (Shopee/Lazada) and return final HTML.

    Requires:
      pip install playwright
      playwright install chromium
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError(
            "Playwright is required for Shopee/Lazada scraping. "
            "Install with 'pip install playwright' and "
            "'playwright install chromium'."
        ) from exc

    logging.info("BROWSER GET %s", url)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(url, wait_until="networkidle", timeout=timeout_ms)

        # --- scroll to trigger lazy-loaded product list ---
        last_height = 0
        for _ in range(10):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
            page.wait_for_timeout(1000)
            new_height = page.evaluate("document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        # small extra wait for JS to render cards
        page.wait_for_timeout(2000)

        html = page.content()

        # optional: dump first page HTML for debugging
        debug_path = os.path.join(os.getcwd(), "debug_shopee_page.html")
        try:
            with open(debug_path, "w", encoding="utf-8") as f:
                f.write(html)
            logging.info("Dumped rendered HTML to %s", debug_path)
        except Exception as e:
            logging.warning("Failed to dump debug HTML: %s", e)

        browser.close()
    return html


def update_page_query(base_url: str, page: int) -> str:
    """Set page=X in query string while preserving other params."""
    p = urlparse(base_url)
    q = parse_qs(p.query)
    q["page"] = [str(page)]
    new_q = urlencode(q, doseq=True)
    return urlunparse(p._replace(query=new_q))

def scrape_shopee_product_via_api(product_url: str, shop_url: str) -> dict | None:
    """
    Call Shopee's item API for a single product URL and map it
    into our internal product dict.
    """
    ids = parse_shopee_ids(product_url)
    if not ids:
        logging.warning("Shopee API: could not parse shopid/itemid from %s", product_url)
        return None

    shopid, itemid = ids
    params = {"itemid": itemid, "shopid": shopid}

    # Use a clean header set tuned for JSON
    api_headers = {
        "User-Agent": HEADERS["User-Agent"],
        "Accept": "application/json, text/plain, */*",
        "Referer": product_url,
    }

    logging.info("Shopee API GET %s params=%s", SHOPEE_ITEM_API, params)
    resp = None
    try:
        resp = SESSION.get(
            SHOPEE_ITEM_API,
            params=params,
            headers={"Referer": product_url, "User-Agent": HEADERS["User-Agent"]},
            timeout=20,
        )
        logging.info("Shopee API status=%s url=%s", resp.status_code, resp.url)
        debug_path = os.path.join(os.getcwd(), "debug_shopee_api.txt")
        with open(debug_path, "w", encoding="utf-8") as f:
            f.write(resp.text[:5000])
        resp.raise_for_status()
        payload = resp.json()
    except Exception as e:
        logging.warning("Shopee API failed for %s: %s", product_url, e)
        # dump body so you can inspect what Shopee is actually sending back
        try:
            if resp is not None:
                debug_path = os.path.join(os.getcwd(), "debug_shopee_api.txt")
                with open(debug_path, "w", encoding="utf-8") as f:
                    f.write(resp.text[:4000])
                logging.info("Dumped Shopee API body to %s", debug_path)
        except Exception:
            pass
        return None

    item = payload.get("data") or payload.get("item")
    if not item:
        logging.warning("Shopee API: no 'data' in response for %s", product_url)
        return None

    # --- ratings ---
    rating_info = item.get("item_rating") or {}
    rating_star = rating_info.get("rating_star", "")
    rc = rating_info.get("rating_count") or []
    rating_count = rc[0] if rc else ""

    # --- variants / shades ---
    variant_parts = []
    for v in item.get("tier_variations", []) or []:
        name = (v.get("name") or "").strip()
        opts = ", ".join(v.get("options") or [])
        if name and opts:
            variant_parts.append(f"{name}: {opts}")
        elif opts:
            variant_parts.append(opts)
    variants_str = " | ".join(variant_parts)

    # --- images ---
    image_codes = item.get("images") or []
    image_urls = [f"https://cf.shopee.ph/file/{code}" for code in image_codes if code]

    # Shopee prices are stored as integer * 100000
    def normalize_price(raw):
        if raw in (None, ""):
            return ""
        try:
            return raw / 100000.0
        except Exception:
            return raw

    price = normalize_price(item.get("price"))

    prod = {
        "platform": "shopee",
        "shop_url": shop_url,
        "product_url": product_url,
        "name": (item.get("name") or "").strip(),
        "description": (item.get("description") or "").strip(),
        "price": price,
        "currency": item.get("currency", ""),
        "availability": "active" if item.get("status") == 1 else item.get("item_status", ""),
        "sku": item.get("itemid", ""),
        "brand": item.get("brand", ""),
        "category": item.get("catid", ""),
        "tags": "",
        "variants": variants_str,
        "rating": rating_star,
        "rating_count": rating_count,
        "image_urls": image_urls,
        "raw_jsonld": json.dumps(item, ensure_ascii=False),
    }
    return prod

# -------------- product URL discovery --------------


def discover_product_links_shopee(shop_url: str, max_pages: int = 10) -> list[str]:
    seen = set()
    product_urls: list[str] = []

    for page in range(max_pages):
        url = update_page_query(shop_url, page) if "page=" in shop_url or page > 0 else shop_url
        try:
            html = fetch_rendered_html_with_browser(url)
        except Exception as e:
            logging.warning("Shopee page %s failed: %s", url, e)
            break

        soup = BeautifulSoup(html, "html.parser")

        all_links = soup.find_all("a", href=True)
        logging.info("Page %s: total <a href> count = %d", page, len(all_links))

        if page == 0:
            hrefs_sample = [a["href"] for a in all_links[:50]]
            logging.info("Sample hrefs on page 0: %s", hrefs_sample)

        found_this_page = 0
        for a in all_links:
            href = a["href"]
            full = urljoin(url, href)
            if full in seen:
                continue
            if SHOPEE_PRODUCT_RE.search(href):
                seen.add(full)
                product_urls.append(full)
                found_this_page += 1

        logging.info("Shopee page %s → %d product links", page, found_this_page)
        if found_this_page == 0 and page > 0:
            break

        time.sleep(1.0)

    return product_urls


def discover_product_links_lazada(shop_url: str, max_pages: int = 10) -> list[str]:
    seen = set()
    product_urls: list[str] = []

    for page in range(1, max_pages + 1):
        url = update_page_query(shop_url, page)
        try:
            # Lazada sometimes works with plain HTML, but Playwright is safer
            html = fetch_rendered_html_with_browser(url)
        except Exception as e:
            logging.warning("Lazada page %s failed: %s", url, e)
            break

        soup = BeautifulSoup(html, "html.parser")
        found_this_page = 0

        for a in soup.find_all("a", href=True):
            href = a["href"]
            full = urljoin(url, href)
            if full in seen:
                continue
            if LAZADA_PRODUCT_RE.search(href):
                seen.add(full)
                product_urls.append(full)
                found_this_page += 1

        logging.info("Lazada page %s → %d product links", page, found_this_page)
        if found_this_page == 0 and page > 1:
            break

        time.sleep(1.0)

    return product_urls


def discover_product_links(shop_url: str, max_pages: int = 10) -> list[str]:
    platform = get_platform(shop_url)
    if platform == "shopee":
        return discover_product_links_shopee(shop_url, max_pages=max_pages)
    if platform == "lazada":
        return discover_product_links_lazada(shop_url, max_pages=max_pages)
    raise ValueError(f"Unsupported platform for URL: {shop_url}")


# -------------- product detail extraction --------------


def _find_jsonld_product(soup: BeautifulSoup):
    """Return first JSON-LD object with @type Product, if present."""
    candidates = []

    for script in soup.find_all("script", type="application/ld+json"):
        raw = script.string or script.get_text(strip=True)
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            continue

        def collect(obj):
            if isinstance(obj, dict):
                t = obj.get("@type") or obj.get("@type".lower())
                if isinstance(t, list):
                    t_list = [str(x).lower() for x in t]
                else:
                    t_list = [str(t).lower()] if t else []
                if any("product" in tt for tt in t_list):
                    candidates.append(obj)
            elif isinstance(obj, list):
                for item in obj:
                    collect(item)

        collect(data)

    return candidates[0] if candidates else None


def extract_from_jsonld(prod: dict, url: str) -> dict:
    offers = prod.get("offers") or {}
    if isinstance(offers, list):
        offers = offers[0] if offers else {}

    images = prod.get("image") or []
    if isinstance(images, str):
        images = [images]

    brand = prod.get("brand")
    if isinstance(brand, dict):
        brand_name = brand.get("name", "")
    else:
        brand_name = brand or ""

    tags = prod.get("keywords", "")
    if isinstance(tags, list):
        tags = ", ".join(tags)

    variants = []
    for key in ("color", "size", "pattern", "material"):
        if key in prod:
            variants.append(f"{key}={prod[key]}")
    variants_str = "; ".join(variants)

    return {
        "platform": "",
        "shop_url": "",
        "product_url": url,
        "name": prod.get("name", "").strip(),
        "description": prod.get("description", "").strip(),
        "price": offers.get("price", ""),
        "currency": offers.get("priceCurrency", ""),
        "availability": offers.get("availability", ""),
        "sku": prod.get("sku", "") or prod.get("mpn", ""),
        "brand": brand_name,
        "category": prod.get("category", ""),
        "tags": tags,
        "variants": variants_str,
        "image_urls": images,
        "raw_jsonld": json.dumps(prod, ensure_ascii=False),
    }


def extract_from_meta(soup: BeautifulSoup, url: str) -> dict:
    def meta(name=None, prop=None, default=""):
        if name:
            tag = soup.find("meta", attrs={"name": name})
        else:
            tag = soup.find("meta", attrs={"property": prop})
        return tag.get("content", "").strip() if tag and tag.get("content") else default

    images = []
    og_image = meta(prop="og:image", default="")
    if og_image:
        images.append(og_image)

    return {
        "platform": "",
        "shop_url": "",
        "product_url": url,
        "name": meta(prop="og:title", default=soup.title.get_text(strip=True) if soup.title else ""),
        "description": meta(prop="og:description", default=meta(name="description", default="")),
        "price": meta(prop="product:price:amount", default=""),
        "currency": meta(prop="product:price:currency", default=""),
        "availability": meta(prop="product:availability", default=""),
        "sku": meta(name="sku", default=""),
        "brand": meta(name="brand", default=""),
        "category": "",
        "tags": meta(name="keywords", default=""),
        "variants": "",
        "image_urls": images,
        "raw_jsonld": "",
    }


def scrape_product(url: str, platform: str, shop_url: str) -> dict | None:
    # Shopee: prefer API, fall back to HTML if API fails
    if platform == "shopee" and is_shopee_product_url(url):
        prod = scrape_shopee_product_via_api(url, shop_url)
        if prod:
            return prod
        logging.info("Falling back to HTML scraping for %s", url)

    try:
        if platform == "shopee":
            html = fetch_rendered_html_with_browser(url)
            resp_url = url
        else:
            resp = fetch(url)
            html = resp.text
            resp_url = resp.url
    except Exception as e:
        logging.warning("Failed product %s: %s", url, e)
        return None

    soup = BeautifulSoup(html, "html.parser")

    # --- Shopee block / homepage detection ---
    page_title = soup.title.get_text(strip=True) if soup.title else ""
    if (platform == "shopee"
            and page_title.startswith("Shopee Philippines | Shop Online")):
        logging.warning(
            "Shopee returned generic homepage / blocked page for %s. "
            "Skipping product.", url
        )
        return None
    # -----------------------------------------

    prod_jsonld = _find_jsonld_product(soup)

    if prod_jsonld:
        data = extract_from_jsonld(prod_jsonld, resp_url)
    else:
        data = extract_from_meta(soup, resp_url)

    data.setdefault("rating", "")
    data.setdefault("rating_count", "")
    data["platform"] = platform
    data["shop_url"] = shop_url
    return data

# -------------- Excel + image ZIP --------------


def save_products_excel(products: list[dict], out_path: str) -> str:
    if not products:
        raise ValueError("No products to write.")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "products"

    headers = [
        "platform",
        "shop_url",
        "product_url",
        "name",
        "description",
        "price",
        "currency",
        "availability",
        "sku",
        "brand",
        "category",
        "tags",
        "variants",
        "rating",
        "rating_count",
        "image_urls",
        "raw_jsonld",
    ]
    ws.append(headers)

    for prod in products:
        row = []
        for h in headers:
            v = prod.get(h, "")
            if h == "image_urls" and isinstance(v, list):
                v = ", ".join(v)
            row.append(v)
        ws.append(row)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 30

    wb.save(out_path)
    logging.info("Saved product Excel → %s", out_path)
    return out_path


def download_images_and_zip(products: list[dict], out_dir: str, zip_name: str = "images.zip") -> str:
    img_root = os.path.join(out_dir, "images")
    os.makedirs(img_root, exist_ok=True)

    downloaded_files: list[str] = []

    for prod in products:
        base = slugify(prod.get("name") or prod.get("sku") or "product")
        for idx, img_url in enumerate(prod.get("image_urls", []), start=1):
            if not img_url:
                continue
            try:
                r = SESSION.get(img_url, timeout=30)
                r.raise_for_status()
            except Exception as e:
                logging.warning("Image download failed %s: %s", img_url, e)
                continue

            ext = os.path.splitext(urlparse(img_url).path)[1] or ".jpg"
            fname = f"{base}_{idx}{ext}"
            fpath = os.path.join(img_root, fname)
            with open(fpath, "wb") as f:
                f.write(r.content)
            downloaded_files.append(fpath)

    zip_path = os.path.join(out_dir, zip_name)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for fp in downloaded_files:
            arc = os.path.relpath(fp, start=out_dir)
            z.write(fp, arcname=arc)

    logging.info("Images zipped → %s", zip_path)
    return zip_path


# -------------- public entrypoint --------------


def scrape_shop(
    shop_url: str,
    out_dir: str = "shop_output",
    max_pages: int = 10,
    include_excel: bool = True,
    include_images: bool = True,
    manual_product_urls: list[str] | None = None,  # NEW
) -> tuple[str | None, str | None]:
    """
    Returns (excel_path | None, images_zip_path | None)
    based on include_excel / include_images flags.
    """

    platform = get_platform(shop_url)
    if platform == "unknown":
        raise ValueError(f"Cannot detect platform from URL: {shop_url}")

    logging.info("Platform: %s", platform)

    # --- 1) Manual URL mode wins ---
    if manual_product_urls:
        product_links = manual_product_urls
        logging.info("Using %d manually supplied product URLs", len(product_links))

    # --- 2) If shop_url itself looks like a product, treat it as single product ---
    elif platform == "shopee" and is_shopee_product_url(shop_url):
        product_links = [shop_url]
        logging.info("Shop URL looks like a single Shopee product. Scraping it directly.")

    # --- 3) Fallback: try auto-discovery from shop listing page ---
    else:
        product_links = discover_product_links(shop_url, max_pages=max_pages)
        logging.info("Discovered %d product URLs", len(product_links))

    products: list[dict] = []
    for url in product_links:
        p = scrape_product(url, platform=platform, shop_url=shop_url)
        if p:
            products.append(p)
        time.sleep(0.8)  # be polite

    if not products:
        raise ValueError("No products scraped.")

    os.makedirs(out_dir, exist_ok=True)

    excel_path: str | None = None
    images_zip_path: str | None = None

    if include_excel:
        excel_path = os.path.join(out_dir, "products.xlsx")
        excel_path = save_products_excel(products, excel_path)

    if include_images:
        images_zip_path = download_images_and_zip(products, out_dir=out_dir)

    return excel_path, images_zip_path


if __name__ == "__main__":
    # Example usage for your Noyona shop:
    url = "https://shopee.ph/noyona_official?entryPoint=ShopBySearch&searchKeyword=noyona"
    excel, img_zip = scrape_shop(url, out_dir="noyona_shopee", max_pages=10)
    print("Excel:", excel)
    print("Images ZIP:", img_zip)

# crawler_excel.py
import os, re, time, random, logging, hashlib, zipfile, pathlib
from email.utils import parsedate_to_datetime
from datetime import datetime, timezone
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse, urlunparse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import deque, defaultdict
import textstat

logging.basicConfig(level=logging.INFO)

# ----------------------------- HTTP SESSION -----------------------------

HEADERS = {
    # Realistic desktop Chrome UA + standard browsery headers
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Upgrade-Insecure-Requests": "1",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}
SESSION = requests.Session()
SESSION.headers.update(HEADERS)
SESSION.max_redirects = 5  # defensive

# Per-host last-request timestamps for polite pacing
_LAST_REQ_TS = defaultdict(lambda: 0.0)

def _sleep_for_rate_limit(url: str, rpm: int, jitter_ratio: float = 0.25):
    """Simple per-host pacing: ensures ~rpm requests per minute per host, with jitter."""
    host = urlparse(url).netloc
    min_interval = 60.0 / max(1, rpm)
    now = time.monotonic()
    elapsed = now - _LAST_REQ_TS[host]
    if elapsed < min_interval:
        # jitter up to 25% of min_interval (default) to avoid lockstep
        jitter = random.uniform(0, min_interval * jitter_ratio)
        time.sleep((min_interval - elapsed) + jitter)
    _LAST_REQ_TS[host] = time.monotonic()

def _parse_retry_after(header_val: str) -> float | None:
    """Return seconds to wait from Retry-After header (either seconds or HTTP-date)."""
    if not header_val:
        return None
    header_val = header_val.strip()
    if header_val.isdigit():
        return max(0.0, float(header_val))
    try:
        dt = parsedate_to_datetime(header_val)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return max(0.0, (dt - datetime.now(timezone.utc)).total_seconds())
    except Exception:
        return None

def _base_host(h: str) -> str:
    h = (h or "").lower()
    return h[4:] if h.startswith("www.") else h

CANON_HOST = None  # pinned canonical host set in crawl_pages()

def same_domain(url, root):
    u = _base_host(urlparse(url).hostname or "")
    r = _base_host(urlparse(root).hostname or "")
    return u == r or u.endswith("." + r)

def _pin_host(u: str) -> str:
    """Force links to use the original start_url host (prevents www <-> apex flip 404)."""
    if not CANON_HOST:
        return u
    p = urlparse(u)
    if _base_host(p.netloc) == _base_host(CANON_HOST) and p.netloc != CANON_HOST:
        p = p._replace(netloc=CANON_HOST)
    return urlunparse(p)

# Skip assets / utility URLs so we don't waste requests
_ASSET_EXT_RE = re.compile(
    r"\.(?:png|jpe?g|gif|webp|svg|ico|bmp|mp4|webm|mp3|wav|pdf|docx?|xlsx?|pptx?|zip|rar|7z)(?:\?.*)?$",
    re.I
)
def _is_crawlable_http_url(u: str) -> bool:
    if not u or not u.startswith(("http://", "https://")):
        return False
    if _ASSET_EXT_RE.search(urlparse(u).path or ""):
        return False
    p = urlparse(u)
    if p.scheme not in ("http", "https"):
        return False
    # Skip common CDN/utility paths that show up as links
    if "/cdn-cgi/" in p.path:
        return False
    return True

def _normalize(u: str) -> str:
    p = urlparse(u)
    path = re.sub(r"/{2,}", "/", p.path or "/")
    # Only add trailing slash if path looks like a directory (no file extension)
    if path != "/" and not path.endswith("/"):
        if not re.search(r"/[^/]+\.[A-Za-z0-9]{1,8}$", path):
            path += "/"
    p = p._replace(path=path)
    return _pin_host(urlunparse(p))

def _get_robots_crawl_delay(start_url: str) -> float | None:
    """Very small robots.txt parser to honor Crawl-delay if present."""
    try:
        robots_url = urljoin(start_url, "/robots.txt")
        _sleep_for_rate_limit(robots_url, rpm=30)  # don't hammer robots either
        r = SESSION.get(robots_url, timeout=10)
        if not r.ok or not r.text:
            return None
        ua = None
        delay_for_star = None
        for raw in r.text.splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):  # comments
                continue
            if line.lower().startswith("user-agent:"):
                ua = line.split(":", 1)[1].strip().lower()
                continue
            if line.lower().startswith("crawl-delay:"):
                try:
                    val = float(line.split(":", 1)[1].strip())
                except Exception:
                    continue
                if ua in (None, "", "*"):
                    delay_for_star = val
                # We send Chrome UA, but most robots files only specify "*"
        return delay_for_star
    except Exception:
        return None

def fetch(url: str, timeout: int = 15, max_hops: int = 10, max_retries: int = 4,
          rate_limit_rpm: int = 12, referer: str | None = None):
    cur = url
    hops = 0
    tries = 0
    chain = []  # collect (status, from, to)

    while True:
        _sleep_for_rate_limit(cur, rpm=rate_limit_rpm)
        headers = SESSION.headers.copy()
        if referer:
            headers["Referer"] = referer

        r = SESSION.get(cur, timeout=timeout, allow_redirects=False, headers=headers)

        if 300 <= r.status_code < 400 and hops < max_hops:
            loc = r.headers.get("Location")
            if not loc:
                # attach chain before returning
                r._redirect_chain = chain
                r._redirected = bool(chain)
                return r
            nxt = urljoin(cur, loc)
            chain.append((r.status_code, cur, nxt))
            referer = cur
            cur = nxt
            hops += 1
            continue

        if r.status_code in (429, 503):
            tries += 1
            retry_after = _parse_retry_after(r.headers.get("Retry-After", ""))
            wait = (max(0.5, retry_after) if retry_after is not None else min(8.0, (2 ** (tries - 1)))) + random.uniform(0.2, 0.8)
            time.sleep(wait)
            if tries < max_retries:
                continue

        # attach chain before returning
        r._redirect_chain = chain
        r._redirected = bool(chain)
        return r

# ----------------------------- BLOG/TYPE -----------------------------

BLOG_HINTS = [
    "/blog", "/article", "/news", "/posts", "/stories", "/insights",
    "/definition", "/definitions", "/review", "/reviews",
    "/how-to", "/how-tos", "/howto", "/guide", "/guides", "/tutorial",
    "/category"
]
DATE_RE = re.compile(r"/\d{4}/\d{2}/")  # WP yyyy/mm

def is_blog_path(path: str) -> bool:
    p = path.lower()
    return DATE_RE.search(p) is not None or any(h in p for h in BLOG_HINTS)

def clean(text):
    return re.sub(r"\s+", " ", text or "").strip()

def get_output_directory(url, base_out_dir="output_excels"):
    parsed = urlparse(url)
    path = parsed.path.lower()
    netloc = parsed.netloc.lower()
    query = parsed.query.lower()

    if "/ko/" in path or netloc.startswith("ko.") or "lang=ko" in query:
        lang_folder = "ko"
    elif "/ja/" in path or netloc.startswith("ja.") or "lang=ja" in query:
        lang_folder = "ja"
    else:
        lang_folder = "default"

    type_folder = "blog" if is_blog_path(path) else "landing-pages"

    out_dir = os.path.join(base_out_dir, lang_folder, type_folder)
    return out_dir, os.path.join(out_dir, "individual")

# ----------------------------- SCRAPERS -----------------------------

def scrape_html_content(soup, base_url, pattern=None):
    def match(text): return not pattern or (text and pattern.search(text.lower()))
    def match_alt(alt): return not pattern or (alt and pattern.search(alt.lower()))

    images = []
    for i in soup.find_all("img"):
        src = i.get("src")
        if not src:
            continue
        images.append({"src": urljoin(base_url, src), "alt": clean(i.get("alt"))})
    if pattern:
        images = [d for d in images if match_alt(d["alt"])]

    return {
        "h1": [clean(t.get_text()) for t in soup.find_all("h1") if match(t.get_text())],
        "h2": [clean(t.get_text()) for t in soup.find_all("h2") if match(t.get_text())],
        "h3": [clean(t.get_text()) for t in soup.find_all("h3") if match(t.get_text())],
        "h4": [clean(t.get_text()) for t in soup.find_all("h4") if match(t.get_text())],
        "h5": [clean(t.get_text()) for t in soup.find_all("h5") if match(t.get_text())],
        "h6": [clean(t.get_text()) for t in soup.find_all("h6") if match(t.get_text())],
        "p":  [clean(p.get_text()) for p in soup.find_all("p")  if match(p.get_text())],
        "images": images,
    }

def scrape_url_info(resp, soup):
    title = soup.title.get_text(strip=True) if soup.title else ""
    metas = {m.get("name", "").lower(): m.get("content", "") for m in soup.find_all("meta") if m.get("name")}
    canonical = [l.get("href") for l in soup.find_all("link", rel="canonical")]
    robots = metas.get("robots", "")
    desc = metas.get("description", "")
    words = len(soup.get_text().split())
    sentences = max(soup.get_text().count(".") or 1, 1)

    h2_tags = [h2.get_text(strip=True) for h2 in soup.find_all("h2")]
    h2_1 = h2_tags[0] if len(h2_tags) > 0 else ""
    h2_2 = h2_tags[1] if len(h2_tags) > 1 else ""

    meta_refresh = ""
    meta_refresh_tag = soup.find("meta", attrs={"http-equiv": re.compile("refresh", re.I)})
    if meta_refresh_tag:
        meta_refresh = meta_refresh_tag.get("content", "")

    rel_next = soup.find("link", rel="next")
    rel_prev = soup.find("link", rel="prev")
    amp_link = soup.find("link", rel="amphtml")

    all_links = [a.get("href") for a in soup.find_all("a", href=True)]
    base = urlparse(resp.url).netloc
    absolute_links = [urljoin(resp.url, a) for a in all_links]
    internal_links = [u for u in absolute_links if urlparse(u).netloc == base]
    external_links = [u for u in absolute_links if urlparse(u).netloc != base]

    html_length = len(resp.text)
    visible_text = len(soup.get_text())
    text_ratio = round((visible_text / html_length) * 100, 2) if html_length else 0

    try:
        flesch = round(textstat.flesch_reading_ease(soup.get_text()), 2)
        readability_label = (
            "Very Easy" if flesch > 90 else
            "Easy" if flesch > 80 else
            "Fairly Easy" if flesch > 70 else
            "Standard" if flesch > 60 else
            "Fairly Difficult" if flesch > 50 else
            "Difficult" if flesch > 30 else "Very Difficult"
        )
    except Exception:
        flesch, readability_label = "", ""

    page_size = len(resp.content)
    co2_mg = round(page_size / 1000 * 0.2, 2)
    carbon_rating = ("A" if co2_mg < 100 else "B" if co2_mg < 200 else
                     "C" if co2_mg < 300 else "D" if co2_mg < 400 else "E")

    mobile_alt = soup.find("link", rel="alternate", media=re.compile("mobile", re.I))
    semantic_similarity = len(set([w.lower() for w in title.split()]) & set([w.lower() for w in desc.split()]))

    status_map = {200: "OK", 301: "Moved Permanently", 302: "Moved Temporarily",
                  404: "Not Found", 429: "Too Many Requests"}

    if 300 <= resp.status_code < 400:
        index_status = "Redirected"
    elif 400 <= resp.status_code < 500:
        index_status = "Client Error"
    elif 500 <= resp.status_code < 600:
        index_status = "Server Error"
    else:
        index_status = "OK"

    path_only = urlparse(resp.url).path.strip("/")
    folder_depth = path_only.count("/") + (1 if path_only else 0)
    crawl_depth = folder_depth
    redirect_chain = getattr(resp, "_redirect_chain", [])

    # --- FIX: authoritative indexability ---
    indexability = "Indexable" if "noindex" not in robots else "Non-Indexable"
    if 300 <= resp.status_code < 400 or redirect_chain:
        indexability = "Non-Indexable"
    # ---------------------------------------

    return {
        "URL": resp.url,
        "Content Type": resp.headers.get("Content-Type", ""),
        "Status Code": resp.status_code,
        "Status": status_map.get(resp.status_code, "Other"),
        "Indexability": indexability,
        "Indexability Status": index_status,
        "Title 1": title,
        "Title 1 Length": len(title),
        "Title 1 Pixel Width": len(title) * 9,
        "Meta Description 1": desc,
        "Meta Description 1 Length": len(desc),
        "Meta Description 1 Pixel Width": len(desc) * 8,
        "Meta Description 2": metas.get("og:description", ""),
        "Meta Description 2 Length": len(metas.get("og:description", "")),
        "Meta Description 2 Pixel Width": len(metas.get("og:description", "")) * 8,
        "Meta Keywords 1": metas.get("keywords", ""),
        "Meta Keywords 1 Length": len(metas.get("keywords", "")),
        "H1-1": (soup.h1.get_text(strip=True) if soup.h1 else ""),
        "H1-1 Length": (len(soup.h1.get_text()) if soup.h1 else 0),
        "H2-1": h2_1,
        "H2-1 Length": len(h2_1),
        "H2-2": h2_2,
        "H2-2 Length": len(h2_2),
        "Meta Robots 1": robots,
        "X-Robots-Tag 1": resp.headers.get("X-Robots-Tag", ""),
        "Meta Refresh 1": meta_refresh,
        "Canonical Link Element 1": canonical[0] if canonical else "",
        "Canonical Link Element 2": canonical[1] if len(canonical) > 1 else "",
        'rel="next" 1': rel_next["href"] if rel_next else "",
        'rel="prev" 1': rel_prev["href"] if rel_prev else "",
        'HTTP rel="next" 1': "",
        'HTTP rel="prev" 1': "",
        "amphtml Link Element": amp_link["href"] if amp_link else "",
        "Size (bytes)": page_size,
        "Transferred (bytes)": page_size,
        "Total Transferred (bytes)": page_size,
        "CO2 (mg)": co2_mg,
        "Carbon Rating": carbon_rating,
        "Word Count": words,
        "Sentence Count": sentences,
        "Average Words Per Sentence": round(words / sentences, 2),
        "Flesch Reading Ease Score": flesch,
        "Readability": readability_label,
        "Text Ratio": text_ratio,
        "Crawl Depth": crawl_depth,
        "Folder Depth": folder_depth,
        "Link Score": len(all_links),
        "Inlinks Unique": len(internal_links),
        "Inlinks Unique JS": 0,
        "Inlinks % of Total": round((len(internal_links) / len(all_links)) * 100, 2) if all_links else 0,
        "Outlinks": len(all_links),
        "Unique Outlinks": len(set(all_links)),
        "Unique JS Outlinks": 0,
        "External Outlinks": len(external_links),
        "Unique External Outlinks": len(set(external_links)),
        "Unique External JS Outlinks": 0,
        "Closest Near Duplicate Match": "",
        "No. Near Duplicates": "",
        "Spelling Errors": 0,
        "Grammar Errors": 0,
        "Hash": hashlib.md5(resp.content).hexdigest(),
        "Response Time": resp.elapsed.total_seconds(),
        "Last Modified": resp.headers.get("Last-Modified", ""),
        "Redirect URL": resp.headers.get("Location", ""),
        "Redirect Type": resp.status_code if 300 <= resp.status_code < 400 else "",
        "Cookies Language": resp.headers.get("Content-Language", ""),
        "HTTP Version": getattr(resp.raw, "version", ""),
        "Mobile Alternate Link": mobile_alt["href"] if mobile_alt else "",
        "Closest Semantically Similar Address": canonical[0] if canonical else "",
        "Semantic Similarity Score": semantic_similarity,
        "No. Semantically Similar": len(set([semantic_similarity])),
        "Semantic Relevance Score": round(semantic_similarity / max(len(title.split()), 1), 2) if title else "",
        "URL Encoded Address": requests.utils.quote(resp.url, safe=""),
        "Crawl Timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),  
        "Final URL": resp.url,
        "Redirected?": 1 if redirect_chain else 0,
        "Redirect Chain": " -> ".join([f"{status}:{src} => {dst}" for status, src, dst in redirect_chain]),
    }

def scrape_performance(resp, soup):
    html_size = len(resp.text)
    img_tags = soup.find_all("img")
    js_tags = soup.find_all("script", src=True)
    css_tags = soup.find_all("link", rel="stylesheet")
    img_count = len(img_tags)

    def parse_width(val):
        if not val:
            return 0
        val = str(val).lower().replace("px", "").replace("%", "").strip()
        try:
            return int(val)
        except ValueError:
            return 0

    large_imgs = [i for i in img_tags if parse_width(i.get("width")) > 2000]

    session = requests.Session()
    session.headers.update(HEADERS)

    # Check for uncompressed images (sample first 5)
    uncompressed_images = 0
    for img in img_tags[:5]:
        src = urljoin(resp.url, img.get("src", ""))
        try:
            r = session.head(src, timeout=10, allow_redirects=True)
            size = int(r.headers.get("Content-Length", 0))
            if size >= 500_000:
                uncompressed_images += 1
        except Exception:
            continue

    # Approximate JS/CSS combined asset size (sample first 5)
    total_asset_size = 0
    checked = 0
    for tag in js_tags + css_tags:
        if checked >= 5:
            break
        src = urljoin(resp.url, tag.get("src") or tag.get("href"))
        try:
            r = session.head(src, timeout=10, allow_redirects=True)
            size = int(r.headers.get("Content-Length", 0))
            total_asset_size += size
            checked += 1
        except Exception:
            continue

    session.close()

    return {
        "Slow Page": html_size > 2_000_000,
        "Large Page Size": html_size,
        "Large Image Size": len(large_imgs),
        "Uncompressed Page": not resp.headers.get("Content-Encoding"),
        "Uncompressed Images": uncompressed_images > 0,
        "Too Many Resources": img_count > 100,
        "Too Many DOM Elements": len(soup.find_all()) > 1500,
        "Excessive HTML Size": html_size > 2_000_000,
        "Excessive JS/CSS Size": total_asset_size > 2_000_000,
    }

def scrape_image_analysis(soup, url, timeout=10):
    data = []
    session = requests.Session()
    session.headers.update(HEADERS)

    for img in soup.find_all("img"):
        src = urljoin(url, img.get("src", ""))
        alt = clean(img.get("alt", ""))
        if urlparse(src).netloc != urlparse(url).netloc:
            continue

        broken, redirected, large_file = False, False, False
        img_size_str = "0 KB"
        try:
            r = session.head(src, allow_redirects=True, timeout=timeout)
            broken = not r.ok
            redirected = len(r.history) > 0
            size = int(r.headers.get("Content-Length", 0))
            if size > 0:
                img_size_str = f"{size / 1_000_000:.2f} MB" if size >= 1_000_000 else f"{size / 1024:.0f} KB"
            large_file = size >= 200_000
        except requests.RequestException:
            broken = True

        entry = {
            "url": src,
            "Missing Alt Text (true false)": alt == "",
            "Broken Image": broken,
            "Large Image File": large_file,
            "size of the image": img_size_str,
            "Image with Redirect": redirected,
            "Image with No Alt Attribute": "alt" not in img.attrs,
            "Image with Empty Alt Text": alt == "",
            "Image with Non-Descriptive Alt Text": len(alt) < 5 and alt != ""
        }
        data.append(entry)

    session.close()
    return data

# ----------------------------- MASTER SCRAPER -----------------------------

def scrape_page(url, referer=None, keywords=None, crawl_types=None, rate_limit_rpm=12):
    logging.info(f"Scraping {url}")
    try:
        resp = fetch(url, rate_limit_rpm=rate_limit_rpm, referer=referer)
        # If a host flip produced 404, retry once pinned
        if resp.status_code == 404 and _base_host(urlparse(resp.url).netloc) != _base_host(urlparse(url).netloc):
            retry = _pin_host(resp.url)
            resp = fetch(retry, rate_limit_rpm=rate_limit_rpm, referer=referer)
        resp.raise_for_status()
    except Exception as e:
        logging.warning(f"Failed to load {url}: {e}")
        return None

    soup = BeautifulSoup(resp.text, "html.parser")
    base = resp.url

    pattern = None
    if keywords:
        pattern = re.compile("|".join([re.escape(k.lower()) for k in keywords]), re.IGNORECASE)

    results = {}
    if "html" in (crawl_types or []):
        results["html"] = scrape_html_content(soup, base, pattern)
    if "url_info" in (crawl_types or []):
        results["url_info"] = scrape_url_info(resp, soup)
    if "performance" in (crawl_types or []):
        results["performance"] = scrape_performance(resp, soup)
    if "images" in (crawl_types or []):
        results["images"] = scrape_image_analysis(soup, base)

    # Extract crawlable links (normalized, absolute, same scheme)
    links = []
    for a in soup.find_all("a", href=True):
        h = urljoin(base, a["href"]).split("#", 1)[0]
        if _is_crawlable_http_url(h):
            links.append(h)
    results["links"] = links
    return results

# ----------------------------- WRITERS -----------------------------

def _apply_header_style(ws):
    from openpyxl.styles import PatternFill, Font
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    for col in range(1, len(ws[1]) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font

def write_excel(page_name, data, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()

    for sheet_name, sheet_data in data.items():
        ws = wb.create_sheet(title=sheet_name[:31])

        # Backward-compat: if images are tuples, convert to dicts
        if sheet_name == "images" and isinstance(sheet_data, list) and sheet_data and isinstance(sheet_data[0], tuple):
            sheet_data = [{"src": t[0], "alt": t[1]} for t in sheet_data]

        if sheet_name == "url_info" and isinstance(sheet_data, dict):
            headers = list(sheet_data.keys())
            ws.append(headers)
            ws.append([str(v) for v in sheet_data.values()])
            _apply_header_style(ws)

        elif isinstance(sheet_data, list):
            if len(sheet_data) > 0 and isinstance(sheet_data[0], dict):
                headers = list(sheet_data[0].keys())
                ws.append(headers)
                _apply_header_style(ws)
                for row in sheet_data:
                    ws.append([row.get(h, "") for h in headers])
            else:
                ws.append(["Value"])
                _apply_header_style(ws)
                for val in sheet_data:
                    ws.append([val])

        elif isinstance(sheet_data, dict):
            ws.append(["Field", "Value"])
            _apply_header_style(ws)
            for k, v in sheet_data.items():
                ws.append([k, str(v)])

        if ws.max_column:
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 35

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    path = os.path.join(out_dir, f"{page_name}.xlsx")
    wb.save(path)
    logging.info(f"Saved {path}")

def write_master_excel(all_data, out_dir):
    if not all_data:
        logging.info(f"Skipping master workbook for {out_dir} (no pages).")
        return

    os.makedirs(out_dir, exist_ok=True)

    from openpyxl.styles import PatternFill, Font
    wb = Workbook()
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)

    sample_page = all_data[0][1] if all_data else {}
    available_sections = [k for k in sample_page.keys() if k not in ["links"]]

    fixed_url_headers = [
        "URL","Content Type","Status Code","Status","Indexability","Indexability Status",
        "Title 1","Title 1 Length","Title 1 Pixel Width",
        "Meta Description 1","Meta Description 1 Length","Meta Description 1 Pixel Width",
        "Meta Description 2","Meta Description 2 Length","Meta Description 2 Pixel Width",
        "Meta Keywords 1","Meta Keywords 1 Length",
        "H1-1","H1-1 Length","H2-1","H2-1 Length","H2-2","H2-2 Length",
        "Meta Robots 1","X-Robots-Tag 1","Meta Refresh 1",
        "Canonical Link Element 1","Canonical Link Element 2",
        'rel="next" 1','rel="prev" 1','HTTP rel="next" 1','HTTP rel="prev" 1',
        "amphtml Link Element","Size (bytes)","Transferred (bytes)","Total Transferred (bytes)",
        "CO2 (mg)","Carbon Rating","Word Count","Sentence Count","Average Words Per Sentence",
        "Flesch Reading Ease Score","Readability","Text Ratio","Crawl Depth","Folder Depth",
        "Link Score","Inlinks Unique","Inlinks Unique JS","Inlinks % of Total","Outlinks",
        "Unique Outlinks","Unique JS Outlinks","External Outlinks","Unique External Outlinks",
        "Unique External JS Outlinks","Closest Near Duplicate Match","No. Near Duplicates",
        "Spelling Errors","Grammar Errors","Hash","Response Time","Last Modified",
        "Redirect URL","Redirect Type","Cookies Language","HTTP Version","Mobile Alternate Link",
        "Closest Semantically Similar Address","Semantic Similarity Score","No. Semantically Similar",
        "Semantic Relevance Score","URL Encoded Address","Crawl Timestamp","Final URL","Redirected?","Redirect Chain"
    ]

    for section in available_sections:
        ws = wb.create_sheet(title=section[:31])

        if section == "url_info":
            headers = ["Page"] + fixed_url_headers
        else:
            headers = ["Page"]
            for _, page_data in all_data:
                section_data = page_data.get(section)
                if section == "images" and isinstance(section_data, list) and section_data and isinstance(section_data[0], tuple):
                    section_data = [{"src": t[0], "alt": t[1]} for t in section_data]
                if isinstance(section_data, dict):
                    headers.extend(h for h in section_data.keys() if h not in headers)
                elif isinstance(section_data, list) and section_data:
                    if isinstance(section_data[0], dict):
                        headers.extend(h for h in section_data[0].keys() if h not in headers)
                    else:
                        if "Value" not in headers:
                            headers.append("Value")

        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        for page_name, page_data in all_data:
            section_data = page_data.get(section)
            if section == "images" and isinstance(section_data, list) and section_data and isinstance(section_data[0], tuple):
                section_data = [{"src": t[0], "alt": t[1]} for t in section_data]

            if isinstance(section_data, dict):
                ws.append([page_name] + [str(section_data.get(h, "")) for h in headers[1:]])
            elif isinstance(section_data, list) and section_data:
                if isinstance(section_data[0], dict):
                    for entry in section_data:
                        ws.append([page_name] + [str(entry.get(h, "")) for h in headers[1:]])
                else:
                    for val in section_data:
                        row = [page_name] + [""] * (len(headers) - 2) + [str(val)]
                        ws.append(row)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    lang_name = os.path.basename(os.path.dirname(out_dir))
    type_name = os.path.basename(out_dir)
    path = os.path.join(out_dir, f"all_{lang_name}_{type_name}.xlsx")

    wb.save(path)
    logging.info(f"✅ Combined workbook saved → {path}")

# ----------------------------- PACKAGING -----------------------------

def zip_output(root_dir, zip_path=None):
    root_dir = os.path.abspath(root_dir)
    if zip_path is None:
        zip_path = root_dir + ".zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for path, _, files in os.walk(root_dir):
            for f in files:
                full = os.path.join(path, f)
                arc = os.path.relpath(full, start=root_dir)
                z.write(full, arcname=arc)
    logging.info(f"📦 Zipped → {zip_path}")
    return zip_path

# ----------------------------- LANGUAGE FILTER -----------------------------

def is_allowed_language(url, root_url, language_filter):
    try:
        p = urlparse(url)
        path, netloc, query = p.path.lower(), p.netloc.lower(), p.query.lower()

        if language_filter == "all":
            return True
        if language_filter == "ko":
            return path.startswith("/ko/") or netloc.startswith("ko.") or "lang=ko" in query
        if language_filter == "ja":
            return path.startswith("/ja/") or netloc.startswith("ja.") or "lang=ja" in query

        blocked = (
            path.startswith(("/ko/", "/ja/", "/zh/")) or
            netloc.startswith(("ko.", "ja.", "zh.")) or
            any(q in query for q in ("lang=ko", "lang=ja", "lang=zh"))
        )
        return not blocked
    except Exception:
        return True

# ----------------------------- MAIN CRAWLER -----------------------------

def crawl_pages(start_urls,
                out_dir="output_excels",
                max_pages=50,
                keyword_filter="",
                language_filter="default",
                crawl_types=None,
                page_scope="both",
                zip_results=False,
                save_individual=True,
                rate_limit_rpm=12,
                obey_robots_delay=True):
    """
    rate_limit_rpm: approx requests per minute per host (polite pacing).
    obey_robots_delay: if robots.txt has Crawl-delay, use the slower of that and rate_limit_rpm.
    """
    if crawl_types is None:
        crawl_types = ["html"]

    # Pin canonical host for redirect stability
    global CANON_HOST
    CANON_HOST = urlparse(start_urls[0]).netloc

    # Robots crawl-delay → convert to rpm (slower wins)
    if obey_robots_delay:
        cd = _get_robots_crawl_delay(start_urls[0])
        if cd and cd > 0:
            robots_rpm = max(1, int(60.0 / cd))
            if robots_rpm < rate_limit_rpm:
                logging.info(f"robots.txt crawl-delay detected ({cd}s). Using ~{robots_rpm} rpm.")
                rate_limit_rpm = robots_rpm

    keywords = [k.strip() for k in keyword_filter.split(",") if k.strip()]
    seen, queue = set(), deque(_normalize(u) for u in start_urls)
    referers = { _normalize(start_urls[0]): None }  # track referer per URL we enqueue
    queued = set(queue)
    root = _normalize(start_urls[0])
    all_data = []

    while queue and len(seen) < max_pages:
        raw = queue.popleft()
        url = _normalize(raw)
        if url in seen or not same_domain(url, root):
            continue
        if not is_allowed_language(url, root, language_filter):
            logging.info(f"Skipping {url} due to language filter ({language_filter})")
            continue

        path_lower = urlparse(url).path.lower()
        is_blog = is_blog_path(path_lower)
        if page_scope == "landing" and is_blog:
            logging.info(f"Skipping blog/article page: {url}")
            continue
        if page_scope == "blog" and not is_blog:
            logging.info(f"Skipping non-blog page: {url}")
            continue

        page_data = scrape_page(url,
                                referer=referers.get(url),
                                keywords=keywords,
                                crawl_types=crawl_types,
                                rate_limit_rpm=rate_limit_rpm)
        if not page_data:
            # if we got throttled too much, take a longer cooldown and continue
            time.sleep(3.0)
            continue

        seen.add(url)

        page_name = urlparse(url).path.strip("/") or "index"
        page_name = page_name.replace("/", "_")[:80]
        structured_out_dir, individual_dir = get_output_directory(url, out_dir)

        logging.info("=" * 60)
        logging.info(f"📂 [LANG+TYPE] → {individual_dir}")
        logging.info(f"🌐 Crawling URL: {url}")
        logging.info("=" * 60)

        try:
            if save_individual:
                write_excel(page_name, page_data, individual_dir)
        except Exception as e:
            logging.error(f"Failed to write Excel for {page_name}: {e}")

        all_data.append((page_name, page_data, structured_out_dir))

        # Enqueue children
        for link in page_data.get("links", []):
            link = _normalize(link)
            if link in seen or link in queued:
                continue
            if not same_domain(link, root):
                continue
            if not is_allowed_language(link, root, language_filter):
                continue
            if not _is_crawlable_http_url(link):
                continue
            queue.append(link)
            referers[link] = url  # set referer to current page
            queued.add(link)

        # base think-time between pages (extra politeness layer)
        time.sleep(random.uniform(0.4, 1.0))

    grouped = defaultdict(list)
    for page_name, page_data, structured_out_dir in all_data:
        grouped[structured_out_dir].append((page_name, page_data))

    for folder, data_list in grouped.items():
        write_master_excel(data_list, folder)

    logging.info("📂 Verifying generated folder structure...")
    for p in pathlib.Path(out_dir).rglob("*"):
        if p.is_dir():
            logging.info(f"📁 Folder created → {p}")

    logging.info(f"✅ Done: Crawled {len(seen)} pages using modes {crawl_types}")

    if zip_results:
        return zip_output(out_dir)
    return out_dir
